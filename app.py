import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, timedelta
import json
import re
import io
from google.cloud import bigquery
from google.oauth2 import service_account
from openpyxl.styles import PatternFill, Font, Alignment
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# --- PAGE CONFIGURATION ---
st.set_page_config(page_title="Digital Turbine Insight", layout="wide", initial_sidebar_state="expanded")

# ==========================================
# HELPER FUNCTIONS & DIALOGS (For Campaign Dashboard)
# ==========================================
def load_data(uploaded_file):
    try:
        if uploaded_file.name.endswith('.csv'):
            return pd.read_csv(uploaded_file)
        elif uploaded_file.name.endswith(('.xls', '.xlsx')):
            return pd.read_excel(uploaded_file)
    except Exception as e:
        st.error(f"Error reading file {uploaded_file.name}: {e}")
        return None

def clean_columns(df):
    if df is not None and not df.empty:
        if len(df.columns) > 0 and str(df.columns[0]).startswith('string_field_'):
            new_header = df.iloc[0]
            df = df[1:].copy()
            df.columns = new_header

        df.columns = df.columns.astype(str).str.strip().str.replace(r'\s+', ' ', regex=True)

        possible_id_names = ['ID', 'Campaign ID', 'Campaign_ID', 'CID', 'campaign_id', 'id']
        found_id = False
        for col in df.columns:
            if col in possible_id_names:
                df.rename(columns={col: 'ID'}, inplace=True)
                found_id = True
                break

        if not found_id:
            for col in df.columns:
                if 'campaign id' in col.lower() or 'cid' == col.lower():
                    df.rename(columns={col: 'ID'}, inplace=True)
                    break
    return df

def clean_and_convert_float(val):
    if pd.isna(val) or val == "":
        return 0.0
    val_str = str(val).strip()
    val_str = re.sub(r'[$,€£ ]', '', val_str)
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def find_value_smart(row_series, keywords, exclude_keywords=None):
    if row_series.empty: return ""
    cols = row_series.index.tolist()
    for col in cols:
        col_lower = str(col).lower()
        if exclude_keywords:
            if any(exc in col_lower for exc in exclude_keywords):
                continue
        if all(k.lower() in col_lower for k in keywords):
            val = row_series[col]
            if pd.notna(val) and str(val).strip() != "":
                return val
    return ""

STATUS_COLOR_MAP = {
    'delivered in full': {'bg': '38761d', 'text': 'FFFFFF'},
    'over pacing': {'bg': '274e13', 'text': 'FFFFFF'},
    'under pacing': {'bg': 'a61e00', 'text': 'FFFFFF'},
    'on track': {'bg': '93e47d', 'text': '000000'},
    'under delivered': {'bg': 'a61e00', 'text': 'FFFFFF'},
}

def get_status_style(status_val):
    if not isinstance(status_val, str): return None, None
    val_clean = status_val.lower().strip()
    if val_clean in STATUS_COLOR_MAP:
        return STATUS_COLOR_MAP[val_clean]['bg'], STATUS_COLOR_MAP[val_clean]['text']
    return 'f0f0f0', '000000'

def apply_excel_styling(writer, df, sheet_name):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]

    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    metric_fill = PatternFill(start_color="a61c00", end_color="a61c00", fill_type="solid")
    metric_font = Font(color="FFFFFF")

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    col_map = {col: idx + 1 for idx, col in enumerate(df.columns)}
    status_col_idx = col_map.get('Pacing_status')
    metric_cols = ['Pacing', 'Yesterday_Pacing_to_Campaign_Delivery%', 'VCR%', 'Yesterday_pacing%']
    metric_indices = [col_map.get(c) for c in metric_cols if c in col_map]

    for row in worksheet.iter_rows(min_row=2, max_row=len(df) + 1):
        if status_col_idx:
            cell = row[status_col_idx - 1]
            val = str(cell.value) if cell.value else ""
            bg_hex, txt_hex = get_status_style(val)
            if bg_hex:
                cell.fill = PatternFill(start_color=bg_hex, end_color=bg_hex, fill_type="solid")
                cell.font = Font(color=txt_hex)

        for m_idx in metric_indices:
            if m_idx:
                cell = row[m_idx - 1]
                cell.fill = metric_fill
                cell.font = metric_font

def pandas_styler(df):
    def color_coding(row):
        styles = [''] * len(row)
        if 'Pacing_status' in row.index:
            idx = df.columns.get_loc('Pacing_status')
            val = row['Pacing_status']
            bg, txt = get_status_style(val)
            if bg:
                styles[idx] = f'background-color: #{bg}; color: #{txt};'

        metric_cols = ['Pacing', 'Yesterday_Pacing_to_Campaign_Delivery%', 'VCR%', 'Yesterday_pacing%']
        for col in metric_cols:
            if col in row.index:
                idx = df.columns.get_loc(col)
                styles[idx] = 'background-color: #a61c00; color: white;'
        return styles
    return df.style.apply(color_coding, axis=1)

@st.dialog("Status Report Details")
def show_status_popup(status_name, full_df):
    st.write(f"### Campaigns: {status_name}")
    subset_df = full_df[full_df['Pacing_status'] == status_name]
    st.write(f"**Total Records:** {len(subset_df)}")
    st.dataframe(pandas_styler(subset_df), use_container_width=True, height=600)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        subset_df.to_excel(writer, index=False, sheet_name='Subset Report')
        apply_excel_styling(writer, subset_df, 'Subset Report')

    file_name = f"DT_{status_name.replace(' ', '_')}_Report.xlsx"
    st.download_button(label=f"Download '{status_name}' Report", data=buffer.getvalue(), file_name=file_name,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       key=f"dl_popup_{status_name}")

@st.dialog("VCR Report Details")
def show_vcr_popup(label, df_subset):
    st.write(f"### {label}")
    st.write(f"**Total Records:** {len(df_subset)}")
    st.dataframe(pandas_styler(df_subset), use_container_width=True, height=600)

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df_subset.to_excel(writer, index=False, sheet_name='VCR Report')
        apply_excel_styling(writer, df_subset, 'VCR Report')

    file_name = f"DT_{label.replace(' ', '_')}_Report.xlsx"
    st.download_button(label=f"Download {label} Report", data=buffer.getvalue(), file_name=file_name,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key=f"dl_popup_vcr")

def send_report_via_email(df, recipient_email, sender_email, sender_password):
    try:
        current_date = datetime.now().strftime('%Y-%m-%d')
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = recipient_email
        msg['Subject'] = f"Digital Turbine Pacing Report - {current_date}"

        html_body = f"""
        <html><body style="font-family: Arial, sans-serif; color: #333;"><div style="padding: 20px;">
        <h2 style="color: #DA1A42;">Digital Turbine Pacing Report</h2>
        <p>Please find attached the latest report.</p>
        <p><strong>Date:</strong> {current_date}</p>
        </div></body></html>
        """
        msg.attach(MIMEText(html_body, 'html'))

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Pacing Report')
            apply_excel_styling(writer, df, 'Pacing Report')

        part = MIMEBase('application', 'octet-stream')
        part.set_payload(output.getvalue())
        encoders.encode_base64(part)
        filename = f"Pacing_Report_{current_date}.xlsx"
        part.add_header('Content-Disposition', f"attachment; filename= {filename}")
        msg.attach(part)

        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(sender_email, sender_password)
        server.sendmail(sender_email, recipient_email, msg.as_string())
        server.quit()
        return True, "Email sent successfully!"
    except Exception as e:
        return False, str(e)

def get_bq_client(json_file):
    try:
        info = json.load(json_file)
        scopes = ["https://www.googleapis.com/auth/bigquery", "https://www.googleapis.com/auth/drive"]
        credentials = service_account.Credentials.from_service_account_info(info, scopes=scopes)
        return bigquery.Client(credentials=credentials, project=credentials.project_id)
    except Exception as e:
        st.error(f"Failed to authenticate: {e}")
        return None

def get_bq_datasets(client):
    try:
        return [d.dataset_id for d in list(client.list_datasets())]
    except:
        return []

def get_bq_tables(client, dataset_id):
    try:
        dataset_ref = f"{client.project}.{dataset_id}"
        return [t.table_id for t in client.list_tables(dataset_ref)]
    except:
        return []

def load_bq_table(client, dataset_id, table_name):
    try:
        query = f"SELECT * FROM `{client.project}.{dataset_id}.{table_name}`"
        return client.query(query).to_dataframe()
    except Exception as e:
        st.error(f"Error loading table **{table_name}**: {e}")
        return None

# --- SESSION STATE FOR NAVIGATION ---
if 'active_page' not in st.session_state:
    st.session_state.active_page = "Dashboard"

def set_page(page_name):
    st.session_state.active_page = page_name

# --- CUSTOM CSS ---
st.markdown("""
    <style>
        .block-container { padding-top: 5rem; }
        
        /* Sidebar Styling */
        [data-testid="stSidebar"] {
            background-color: #ffffff;
        }

        /* Style for all sidebar buttons */
        .stButton > button {
            width: 100%;
            border-radius: 8px;
            border: none;
            text-align: left;
            background-color: transparent;
            color: #4A5568;
            padding: 8px 12px;
            transition: all 0.3s;
            justify-content: flex-start;
        }

        /* Hover effect */
        .stButton > button:hover {
            background-color: #F7FAFC;
            color: #2D3748;
            border: none;
        }
    </style>
""", unsafe_allow_html=True)

# --- SIDEBAR ---
with st.sidebar:
    st.markdown("### 💠 Digital Turbine Insight\n*Marketing Analytics*")
    st.divider()
    
    # Navigation Groups with Material Icons
    menu_structure = {
        "ANALYTICS": {
            "Campaign Dashboard": "campaign",
            "Dashboard": "dashboard",
            "Operations": "vital_signs",
            "Campaign Performance": "bar_chart",
            "Attribution & MMM": "automation",
            "Incrementality Tests": "experiment",
            "Marketing Funnel": "filter_alt"
        },
        "CHANNELS": {
            "Retail Media": "shopping_cart",
            "Walled Gardens": "lock"
        },
        "INTELLIGENCE": {
            "Creative Intelligence": "palette",
            "Audience Intelligence": "groups",
            "Financial & Budget": "payments",
            "Competitive Intel": "swords"
        },
        "TOOLS": {
            "AI Assistant": "chat_bubble",
            "Settings": "settings"
        }
    }

    for section, items in menu_structure.items():
        st.markdown(f"**{section}**")
        for item, icon_name in items.items():
            is_active = st.session_state.active_page == item
            
            if st.button(
                item, 
                icon=f":material/{icon_name}:", 
                key=f"btn_{item}", 
                on_click=set_page, 
                args=(item,),
                type="secondary" if not is_active else "primary"
            ):
                pass
        st.write("") # Spacer

    st.markdown("<br><br>", unsafe_allow_html=True)
    st.markdown("👤 **John Doe**\n\nAdmin")

# --- GLOBAL DATA & COLOR MAP ---
color_map = {
    'Meta': '#2a9d8f', 
    'Google': '#219ebc', 
    'TikTok': '#7209b7', 
    'Amazon': '#ffb703', 
    'Pinterest': '#e63946', 
    'Snap': '#06d6a0'
}
channels = ['Meta', 'Google', 'TikTok', 'Amazon', 'Pinterest', 'Snap']


# ==========================================
# --- MAIN CONTENT LOGIC ---
# ==========================================

if st.session_state.active_page == "Campaign Dashboard":

    # --- TOP HEADER & FILTERS (Matches theme) ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="cd_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="cd_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="cd_date", label_visibility="collapsed")

    st.title("Campaign Dashboard")
    st.markdown("Digital Turbine Pacing Report and Analytics")

    # --- 1. SESSION STATE INIT ---
    if 'ongoing_df' not in st.session_state: st.session_state['ongoing_df'] = None
    if 'completed_df' not in st.session_state: st.session_state['completed_df'] = None
    if 'tracker_df' not in st.session_state: st.session_state['tracker_df'] = None
    if 'report_df' not in st.session_state: st.session_state['report_df'] = None

    if 'filter_managers' not in st.session_state: st.session_state['filter_managers'] = []
    if 'filter_status' not in st.session_state: st.session_state['filter_status'] = []
    if 'filter_vcr' not in st.session_state: st.session_state['filter_vcr'] = "All"
    if 'filter_campaign_status' not in st.session_state: st.session_state['filter_campaign_status'] = []

    if 'default_start_date' not in st.session_state: st.session_state['default_start_date'] = datetime.now().date()
    if 'default_end_date' not in st.session_state: st.session_state['default_end_date'] = datetime.now().date()

    # --- Sidebar Configuration Data Uploaders (ONLY shown on this page) ---
    st.sidebar.header("Configuration")
    data_source = st.sidebar.selectbox("Select Data Source", ["Manual File Upload", "Google BigQuery"])

    # --- 2. DATA LOADING AREA ---
    with st.container(border=True):
        st.markdown("**Data Configuration**\n\n<span style='color:gray; font-size: 14px;'>Upload files or connect to BigQuery</span>", unsafe_allow_html=True)
        
        if data_source == "Manual File Upload":
            col1, col2, col3 = st.columns(3)
            with col1:
                f1 = st.file_uploader("Upload Ongoing", type=['csv', 'xlsx'])
                if f1: st.session_state['ongoing_df'] = clean_columns(load_data(f1))
            with col2:
                f2 = st.file_uploader("Upload Completed", type=['csv', 'xlsx'])
                if f2: st.session_state['completed_df'] = clean_columns(load_data(f2))
            with col3:
                f3 = st.file_uploader("Upload Tracker", type=['csv', 'xlsx'])
                if f3: st.session_state['tracker_df'] = clean_columns(load_data(f3))

        elif data_source == "Google BigQuery":
            auth_file = st.file_uploader("Upload Service Account JSON", type=['json'])
            if auth_file:
                client = get_bq_client(auth_file)
                if client:
                    datasets = get_bq_datasets(client)
                    if datasets:
                        selected_dataset = st.selectbox("Select Dataset", ["-- Select --"] + datasets)
                        if selected_dataset != "-- Select --":
                            with st.spinner("Fetching tables..."):
                                tables = get_bq_tables(client, selected_dataset)
                                if tables:
                                    st.divider()
                                    c1, c2, c3 = st.columns(3)
                                    t_ongoing = c1.selectbox("Ongoing Table", ["-- Select --"] + tables)
                                    t_completed = c2.selectbox("Completed Table", ["-- Select --"] + tables)
                                    t_tracker = c3.selectbox("Tracker Table", ["-- Select --"] + tables)
                                    if st.button("Load BigQuery Data", type="primary"):
                                        if "-- Select --" in [t_ongoing, t_completed, t_tracker]:
                                            st.warning("Please select all tables.")
                                        else:
                                            with st.spinner("Downloading..."):
                                                odf = load_bq_table(client, selected_dataset, t_ongoing)
                                                cdf = load_bq_table(client, selected_dataset, t_completed)
                                                tdf = load_bq_table(client, selected_dataset, t_tracker)
                                                if any(x is None for x in [odf, cdf, tdf]):
                                                    st.error("Failed to load tables.")
                                                else:
                                                    st.session_state['ongoing_df'] = clean_columns(odf)
                                                    st.session_state['completed_df'] = clean_columns(cdf)
                                                    st.session_state['tracker_df'] = clean_columns(tdf)
                                                    st.success("Data Loaded!")

    # --- 3. PROCESSING ---
    today = pd.to_datetime(datetime.now().date())
    ongoing_df = st.session_state['ongoing_df']
    completed_df = st.session_state['completed_df']
    tracker_df = st.session_state['tracker_df']

    if ongoing_df is not None and completed_df is not None and tracker_df is not None:
        st.write("") # Spacer
        if st.button("Generate Campaign Dashboard", type="primary", use_container_width=True):
            with st.spinner("Processing Data..."):
                try:
                    if 'ID' in ongoing_df.columns:
                        ongoing_df['ID'] = ongoing_df['ID'].astype(str).str.replace(r'\.0$', '', regex=True)
                    if 'ID' in completed_df.columns:
                        completed_df['ID'] = completed_df['ID'].astype(str).str.replace(r'\.0$', '', regex=True)

                    matched_col = None
                    if 'ID' in tracker_df.columns:
                        matched_col = 'ID'
                    elif 'CID' in tracker_df.columns:
                        matched_col = 'CID'
                    elif 'Campaign ID' in tracker_df.columns:
                        matched_col = 'Campaign ID'

                    if matched_col:
                        tracker_df[matched_col] = tracker_df[matched_col].astype(str).str.replace(r'\.0$', '', regex=True)
                        cids = tracker_df[matched_col].unique()
                    else:
                        cids = pd.unique(pd.concat([ongoing_df['ID'], completed_df['ID']], ignore_index=True))

                    results = []
                    for cid in cids:
                        row = {'CID': cid}
                        in_ongoing = cid in ongoing_df['ID'].values if 'ID' in ongoing_df.columns else False
                        in_completed = cid in completed_df['ID'].values if 'ID' in completed_df.columns else False

                        t_match = pd.DataFrame()
                        if matched_col: t_match = tracker_df[tracker_df[matched_col] == cid]
                        tracker_row = t_match.iloc[0] if not t_match.empty else pd.Series()

                        campaign_row = pd.Series()
                        row_ongoing_data = pd.Series()
                        row_completed_data = pd.Series()

                        row['Campaign_name'] = find_value_smart(tracker_row, ['name', 'campaign'])
                        row['Account_Manager'] = find_value_smart(tracker_row, ['manager'])
                        row['Placement_id'] = find_value_smart(tracker_row, ['placement'])
                        row['Analyst'] = find_value_smart(tracker_row, ['analyst'])

                        booked_val = find_value_smart(tracker_row, ['booked', 'limit'])
                        if booked_val == "": booked_val = find_value_smart(tracker_row, ['budget'])
                        if booked_val == "": booked_val = find_value_smart(tracker_row, ['total', 'limit'])
                        if booked_val == "": booked_val = find_value_smart(tracker_row, ['limit'], exclude_keywords=['daily'])
                        if booked_val == "": booked_val = find_value_smart(tracker_row, ['amount'])
                        row['Booked_Limit'] = clean_and_convert_float(booked_val)

                        if in_ongoing or in_completed:
                            source_df = ongoing_df if in_ongoing else completed_df
                            campaign_row = source_df[source_df['ID'] == cid].iloc[0]
                            if in_ongoing: row_ongoing_data = campaign_row
                            if in_completed: row_completed_data = campaign_row
                            row['Price_model'] = find_value_smart(campaign_row, ['price', 'model'])
                            row["CVV'S"] = clean_and_convert_float(campaign_row.get('CVVs', 0))
                            imps = clean_and_convert_float(campaign_row.get('Impressions', 0))
                        else:
                            row['Price_model'] = find_value_smart(tracker_row, ['price', 'model'])
                            row["CVV'S"] = 0.0
                            imps = 0.0

                        raw_start = pd.NaT
                        raw_end = pd.NaT
                        if not tracker_row.empty:
                            val_s = find_value_smart(tracker_row, ['start'])
                            if val_s != "": raw_start = val_s
                            val_e = find_value_smart(tracker_row, ['end'])
                            if val_e != "": raw_end = val_e
                        row['Start_date'] = pd.to_datetime(raw_start, errors='coerce')
                        row['End_date'] = pd.to_datetime(raw_end, errors='coerce')

                        if pd.notna(row['End_date']):
                            if not (in_ongoing or in_completed):
                                row['Campaign_status'] = 'Not Found in Platform'
                            else:
                                row['Campaign_status'] = 'In Progress' if row['End_date'] >= today else 'Completed'
                        else:
                            row['Campaign_status'] = 'Completed' if (in_ongoing or in_completed) else 'Unknown'

                        if not (in_ongoing or in_completed):
                            row['VCR%'] = 0.0
                            row['Remaining_Goal'] = row['Booked_Limit']
                            row['Remaining_Goal_in_Platform'] = 0.0
                            row['No_of_days_remaining'] = 0
                            row['Daily_limit_required'] = 0.0
                            row['Yesterday_pacing%'] = 0.0
                            row['Total_days'] = 0
                            row['Pacing'] = 0.0
                            row['Yesterday_Pacing_to_Campaign_Delivery%'] = 0.0
                            row['Delivered'] = 0.0
                            row['Total_limit_in_platform'] = 0.0
                            row['Daily_Limit_in_platform'] = 0.0
                            row['Yesterday_pacing'] = 0.0
                            row['Pacing_status'] = 'No Data'
                            row['Optimization_Direction'] = "Not Live / Missing in Platform Data"
                            results.append(row)
                            continue

                        def fetch_limit(source_row):
                            val = find_value_smart(source_row, ['total', 'limit'])
                            if val == "": val = find_value_smart(source_row, ['limit'], exclude_keywords=['daily'])
                            if val == "": val = find_value_smart(source_row, ['budget'])
                            return val

                        limit_val = 0
                        if pd.notna(row['End_date']) and row['End_date'] >= today:
                            limit_val = fetch_limit(row_ongoing_data) if in_ongoing else 0
                        else:
                            limit_val = fetch_limit(row_completed_data) if in_completed else 0
                        row['Total_limit_in_platform'] = clean_and_convert_float(limit_val)

                        cutoff_date = today - timedelta(days=1)
                        p_model = str(row['Price_model']).strip().upper()
                        use_ongoing = in_ongoing and (pd.isna(row['End_date']) or row['End_date'] > cutoff_date)
                        active_row = row_ongoing_data if use_ongoing else row_completed_data

                        if p_model == 'CPM':
                            row['Delivered'] = clean_and_convert_float(active_row.get('Impressions', 0))
                        else:
                            row['Delivered'] = clean_and_convert_float(active_row.get('CVVs', 0))

                        row['VCR%'] = (row["CVV'S"] / imps * 100) if imps > 0 else 0
                        row['Remaining_Goal'] = row['Booked_Limit'] - row['Delivered']
                        row['Remaining_Goal_in_Platform'] = row['Total_limit_in_platform'] - row['Delivered']

                        if pd.notna(row['End_date']) and row['End_date'] >= today:
                            row['No_of_days_remaining'] = (row['End_date'] - (today - timedelta(days=1))).days
                        else:
                            row['No_of_days_remaining'] = 0

                        if row['No_of_days_remaining'] > 0:
                            row['Daily_limit_required'] = (row['Total_limit_in_platform'] - row['Delivered']) / row['No_of_days_remaining']
                        else:
                            row['Daily_limit_required'] = 0

                        daily_limit_raw = find_value_smart(campaign_row, ['daily', 'limit'])
                        row['Daily_Limit_in_platform'] = clean_and_convert_float(daily_limit_raw) if in_ongoing else 0

                        if pd.notna(row['End_date']) and row['End_date'] < today:
                            row['Yesterday_pacing'] = 0
                        else:
                            if in_ongoing:
                                yest_val = find_value_smart(row_ongoing_data, ['yesterday', 'delivered'])
                                if yest_val == "": yest_val = find_value_smart(row_ongoing_data, ['yesterday'])
                                row['Yesterday_pacing'] = clean_and_convert_float(yest_val)
                            else:
                                row['Yesterday_pacing'] = 0

                        if row['Daily_limit_required'] > 0:
                            row['Yesterday_pacing%'] = row['Yesterday_pacing'] / row['Daily_limit_required']
                        else:
                            row['Yesterday_pacing%'] = 0

                        if pd.notna(row['Start_date']) and pd.notna(row['End_date']):
                            row['Total_days'] = (row['End_date'] - row['Start_date']).days
                        else:
                            row['Total_days'] = 0

                        delivered = row['Delivered']
                        booked = row['Booked_Limit']
                        end_date = row['End_date']
                        start_date = row['Start_date']

                        if pd.notna(end_date) and pd.notna(start_date) and booked > 0:
                            if end_date <= today:
                                row['Pacing'] = delivered / booked
                            else:
                                total_duration = (end_date - start_date).days
                                elapsed_duration = (today - start_date).days
                                if total_duration > 0:
                                    expected = (booked / total_duration) * elapsed_duration
                                    row['Pacing'] = delivered / expected if expected > 0 else 0
                                else:
                                    row['Pacing'] = 0
                        else:
                            row['Pacing'] = 0

                        numerator = (row['Yesterday_pacing'] * row['No_of_days_remaining']) + row['Delivered']
                        if booked > 0:
                            row['Yesterday_Pacing_to_Campaign_Delivery%'] = numerator / booked
                        else:
                            row['Yesterday_Pacing_to_Campaign_Delivery%'] = 0

                        is_active_future = pd.notna(end_date) and end_date >= today
                        if is_active_future:
                            val = row['Yesterday_Pacing_to_Campaign_Delivery%']
                            if val < 1.0:
                                row['Pacing_status'] = 'Under Pacing'
                            elif val > 1.2:
                                row['Pacing_status'] = 'Over Pacing'
                            else:
                                row['Pacing_status'] = 'On Track'
                        else:
                            val = row['Pacing']
                            if val < 1.0:
                                row['Pacing_status'] = 'Under Delivered'
                            else:
                                row['Pacing_status'] = 'Delivered In Full'

                        W = float(row.get('Daily_Limit_in_platform', 0))
                        X = float(row.get('Yesterday_pacing', 0))
                        V = float(row.get('Daily_limit_required', 0))

                        opt_dir = "Review Initial Status"
                        if W == 0:
                            opt_dir = "No Changes Required" if X >= V else "Optimize to Hit Daily Requirement"
                        elif W > 0:
                            if W < V and X < V:
                                opt_dir = "Optimize to Hit Daily Cap and Change the Daily Cap"
                            elif W >= V and X < V:
                                opt_dir = "Optimize to Hit Daily Limit"
                            elif W >= V:
                                opt_dir = "No Changes Required"
                            elif W < V:
                                opt_dir = "Increase Daily Limit"
                        row['Optimization_Direction'] = opt_dir
                        results.append(row)

                    summary_df = pd.DataFrame(results)
                    if 'Account_Manager' in summary_df.columns:
                        summary_df = summary_df[summary_df['Account_Manager'].astype(str) != 'Not Found']

                    columns_order = [
                        'CID', 'Campaign_name', 'Account_Manager', 'Placement_id', 'Analyst', 'Start_date',
                        'End_date', 'Campaign_status', 'Price_model', 'Pacing', 'Pacing_status',
                        'Yesterday_Pacing_to_Campaign_Delivery%', 'Optimization_Direction',
                        'Booked_Limit', 'Total_limit_in_platform', 'Delivered', "CVV'S", 'VCR%',
                        'Remaining_Goal', 'Remaining_Goal_in_Platform', 'No_of_days_remaining',
                        'Daily_limit_required', 'Daily_Limit_in_platform',
                        'Yesterday_pacing', 'Yesterday_pacing%', 'Total_days'
                    ]

                    for col in columns_order:
                        if col not in summary_df.columns: summary_df[col] = 0

                    summary_df = summary_df[columns_order]
                    summary_df['Start_date'] = pd.to_datetime(summary_df['Start_date']).dt.strftime('%Y-%m-%d')
                    summary_df['End_date'] = pd.to_datetime(summary_df['End_date']).dt.strftime('%Y-%m-%d')

                    summary_df['Pacing'] = (summary_df['Pacing'] * 100).round(2)
                    summary_df['Yesterday_Pacing_to_Campaign_Delivery%'] = (summary_df['Yesterday_Pacing_to_Campaign_Delivery%'] * 100).round(2)
                    summary_df['VCR%'] = summary_df['VCR%'].round(2)
                    summary_df['Yesterday_pacing%'] = (summary_df['Yesterday_pacing%'] * 100).round(2)

                    for col in ['Pacing', 'Yesterday_Pacing_to_Campaign_Delivery%', 'VCR%', 'Yesterday_pacing%']:
                        summary_df[col] = summary_df[col].astype(str) + '%'

                    st.session_state['report_df'] = summary_df

                    temp_start_dates = pd.to_datetime(summary_df['Start_date'], errors='coerce').dropna()
                    if not temp_start_dates.empty:
                        min_start = temp_start_dates.min().date()
                        st.session_state['default_start_date'] = min_start
                    else:
                        st.session_state['default_start_date'] = datetime.now().date()
                    st.session_state['default_end_date'] = datetime.now().date()

                except Exception as e:
                    st.error(f"Processing Error: {e}")

    # --- 4. ADVANCED INTERFACE (Post Processing) ---
    if st.session_state['report_df'] is not None:
        df_to_show = st.session_state['report_df']

        # [A] FILTERS BOX
        with st.container(border=True):
            def reset_filter_state():
                st.session_state['filter_managers'] = []
                st.session_state['filter_status'] = []
                st.session_state['filter_vcr'] = "All"
                st.session_state['filter_campaign_status'] = []

            col_head, col_btn = st.columns([8, 1])
            with col_head:
                st.markdown("**Report Filters**\n\n<span style='color:gray; font-size: 14px;'>Refine your campaign view</span>", unsafe_allow_html=True)
            with col_btn:
                st.button("Reset", on_click=reset_filter_state, type="secondary", use_container_width=True)

            col_f1, col_f2, col_f3, col_f4 = st.columns(4)
            with col_f1:
                all_mgrs = sorted(df_to_show['Account_Manager'].astype(str).unique().tolist())
                sel_mgrs = st.multiselect("Account Manager(s):", all_mgrs, key="filter_managers")
            with col_f2:
                all_stats = sorted(df_to_show['Pacing_status'].astype(str).unique().tolist())
                sel_stats = st.multiselect("Pacing Status(es):", all_stats, key="filter_status")
            with col_f3:
                sel_vcr = st.selectbox("VCR%:", ["All", ">= 90%", "< 90%"], key="filter_vcr")
            with col_f4:
                all_camp_stats = sorted(df_to_show['Campaign_status'].astype(str).unique().tolist())
                sel_camp_stats = st.multiselect("Campaign Status:", all_camp_stats, key="filter_campaign_status")

            col_d1, col_d2 = st.columns(2)
            limit_min = datetime(2000, 1, 1).date()
            limit_max = datetime(2030, 12, 31).date()
            with col_d1:
                start_sel = st.date_input("From:", value=st.session_state['default_start_date'], min_value=limit_min, max_value=limit_max, key="filter_from")
            with col_d2:
                end_sel = st.date_input("To:", value=st.session_state['default_end_date'], min_value=limit_min, max_value=limit_max, key="filter_to")

        filtered_df = df_to_show.copy()

        if sel_mgrs: filtered_df = filtered_df[filtered_df['Account_Manager'].isin(sel_mgrs)]
        if sel_stats: filtered_df = filtered_df[filtered_df['Pacing_status'].isin(sel_stats)]
        if sel_camp_stats: filtered_df = filtered_df[filtered_df['Campaign_status'].isin(sel_camp_stats)]
        if sel_vcr != "All":
            try:
                to_float = lambda x: float(str(x).replace('%', '').strip()) if pd.notna(x) and str(x) != '' else 0.0
                if sel_vcr == ">= 90%": filtered_df = filtered_df[filtered_df['VCR%'].apply(to_float) >= 90.0]
                elif sel_vcr == "< 90%": filtered_df = filtered_df[filtered_df['VCR%'].apply(to_float) < 90.0]
            except: pass
        try:
            s_d = pd.to_datetime(filtered_df['Start_date'], errors='coerce').dt.date
            mask = ((s_d >= start_sel) | s_d.isna()) & ((s_d <= end_sel) | s_d.isna())
            filtered_df = filtered_df[mask]
        except Exception as e:
            st.error(f"Date Filter Error: {e}")

        # [B] QUICK STATS / POP-UP BUTTONS
        st.markdown("**Campaign Status Overview**")
        
        status_counts = filtered_df['Pacing_status'].value_counts()
        
        # Combine Pacing Status and VCR counts into a single row of KPI metrics
        vcr_clean = filtered_df['VCR%'].astype(str).str.replace('%', '', regex=False)
        vcr_numeric = pd.to_numeric(vcr_clean, errors='coerce').fillna(0)
        high_vcr_df = filtered_df[vcr_numeric >= 90]
        low_vcr_df = filtered_df[vcr_numeric < 90]
        
        total_cols = len(status_counts) + 2
        cols = st.columns(total_cols)
        
        for index, (status, count) in enumerate(status_counts.items()):
            with cols[index]:
                with st.container(border=True):
                    st.metric(label=status, value=count)
                    if st.button("View Details", key=f"btn_pop_{status}", use_container_width=True):
                        show_status_popup(status, filtered_df)
                
        with cols[-2]:
            with st.container(border=True):
                st.metric(label="VCR >= 90%", value=len(high_vcr_df))
                if st.button("View Details", key="btn_vcr_high", use_container_width=True):
                    show_vcr_popup("VCR >= 90%", high_vcr_df)
            
        with cols[-1]:
            with st.container(border=True):
                st.metric(label="VCR < 90%", value=len(low_vcr_df))
                if st.button("View Details", key="btn_vcr_low", use_container_width=True):
                    show_vcr_popup("VCR < 90%", low_vcr_df)

        st.write("") # Spacer

        # [C] CHARTS ROW
        if not filtered_df.empty:
            plot_df = filtered_df.copy()
            def clean_numeric_col(x):
                if isinstance(x, str):
                    return float(x.replace('%', '').replace(',', '').strip())
                return float(x)

            try:
                plot_df['Pacing_Num'] = plot_df['Pacing'].apply(clean_numeric_col)
                plot_df['VCR_Num'] = plot_df['VCR%'].apply(clean_numeric_col)
                plot_df['Booked_Limit'] = plot_df['Booked_Limit'].apply(clean_numeric_col)
                plot_df['Delivered'] = plot_df['Delivered'].apply(clean_numeric_col)

                c_chart1, c_chart2 = st.columns(2)

                with c_chart1:
                    with st.container(border=True):
                        st.markdown("**Campaign Count by Status**")
                        fig_status = px.pie(plot_df, names='Pacing_status', hole=0.4)
                        fig_status.update_layout(margin=dict(l=0, r=0, t=20, b=0), height=300)
                        st.plotly_chart(fig_status, use_container_width=True)

                with c_chart2:
                    with st.container(border=True):
                        st.markdown("**Total Budget Value by Status**")
                        status_value = plot_df.groupby('Pacing_status')[['Booked_Limit']].sum().reset_index()
                        color_discrete_map = {
                            'Delivered In Full': '#38761d', 'Over Pacing': '#274e13',
                            'Under Pacing': '#a61e00', 'On Track': '#93e47d', 'Under Delivered': '#a61e00'
                        }
                        fig_value = px.bar(status_value, x='Pacing_status', y='Booked_Limit', text_auto='.2s', color='Pacing_status', color_discrete_map=color_discrete_map)
                        fig_value.update_layout(showlegend=False, margin=dict(l=0, r=0, t=20, b=0), height=300, xaxis_title=None, yaxis_title=None)
                        st.plotly_chart(fig_value, use_container_width=True)

                with st.container(border=True):
                    st.markdown("**Manager Performance**\n\n<span style='color:gray; font-size: 14px;'>Total Budget vs Delivered by Manager</span>", unsafe_allow_html=True)
                    mgr_grp = plot_df.groupby('Account_Manager')[['Booked_Limit', 'Delivered']].sum().reset_index()
                    fig_mgr = px.bar(mgr_grp, x='Account_Manager', y=['Booked_Limit', 'Delivered'], barmode='group')
                    fig_mgr.update_layout(legend_title_text='', legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5), margin=dict(l=0, r=0, t=20, b=0), height=350, xaxis_title=None, yaxis_title=None)
                    st.plotly_chart(fig_mgr, use_container_width=True)

            except Exception as e:
                st.warning(f"Could not generate charts: {e}")

        # [D] MAIN TABLE
        with st.container(border=True):
            st.markdown(f"**Main Table Data**\n\n<span style='color:gray; font-size: 14px;'>{len(filtered_df)} records match current filters</span>", unsafe_allow_html=True)
            st.dataframe(pandas_styler(filtered_df), use_container_width=True, height=450)

        # [E] EXPORT & EMAIL
        with st.container(border=True):
            st.markdown("**Export & Share Report**")
            col_ex1, col_ex2 = st.columns(2)
            
            with col_ex1:
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    filtered_df.to_excel(writer, index=False, sheet_name='Pacing Report')
                    apply_excel_styling(writer, filtered_df, 'Pacing Report')

                st.download_button(
                    label="📥 Download Excel Report",
                    data=buffer.getvalue(),
                    file_name=f"DT_Pacing_Report_{today.strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
                
            with col_ex2:
                with st.popover("✉️ Send via Email", use_container_width=True):
                    recipient = st.text_input("Recipient Email Address")
                    if st.button("Send Email", type="primary"):
                        if not recipient:
                            st.warning("Enter email.")
                        else:
                            try:
                                s_email = st.secrets["SENDER_EMAIL"]
                                s_pass = st.secrets["SENDER_PASSWORD"]
                                with st.spinner("Sending..."):
                                    ok, msg = send_report_via_email(filtered_df, recipient, s_email, s_pass)
                                    if ok: st.success(msg)
                                    else: st.error(msg)
                            except:
                                st.error("Please configure secrets.toml with SENDER_EMAIL and SENDER_PASSWORD.")

    elif (ongoing_df is None or completed_df is None or tracker_df is None):
        st.info("Please load Ongoing, Completed, and Tracker datasets using the 'Data Configuration' block above to proceed.")


# --- DASHBOARD PAGE ---
elif st.session_state.active_page == "Dashboard":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="dash_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="dash_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="dash_date", label_visibility="collapsed")

    st.title("Dashboard")
    st.markdown("Unified marketing performance overview")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        with st.container(border=True):
            st.metric(label="Total Spend", value="$765K", delta="12.5%")
    with m2:
        with st.container(border=True):
            st.metric(label="ROAS", value="4.2x", delta="8.3%")
    with m3:
        with st.container(border=True):
            st.metric(label="Conversions", value="18.5K", delta="-3.2%")
    with m4:
        with st.container(border=True):
            st.metric(label="CPA", value="$41.35", delta="-5.1%", delta_color="inverse")

    # --- DATA PREP ---
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    roas_data = pd.DataFrame({'Month': months, 'ROAS': [3.2, 3.5, 3.1, 3.8, 4.2, 4.0, 4.5, 4.1, 3.9, 4.3, 4.8, 5.1]})
    spend_data = pd.DataFrame({'Channel': channels, 'Spend': [270, 200, 90, 160, 40, 50]}).iloc[::-1] 

    # --- ROW 1 CHARTS ---
    col1, col2 = st.columns([1.2, 1])
    with col1:
        with st.container(border=True):
            st.markdown("**ROAS Trend** \n*Last 12 months*")
            fig_roas = px.line(roas_data, x='Month', y='ROAS')
            fig_roas.update_traces(fill='tozeroy', line_color='#2a9d8f')
            fig_roas.update_layout(margin=dict(l=0, r=0, t=20, b=0), height=300, yaxis_range=[0, 8])
            st.plotly_chart(fig_roas, use_container_width=True)

    with col2:
        with st.container(border=True):
            st.markdown("**Spend by Channel** \n*Current period*")
            fig_spend = px.bar(spend_data, x='Spend', y='Channel', orientation='h', color='Channel', color_discrete_map=color_map)
            fig_spend.update_layout(showlegend=False, margin=dict(l=0, r=0, t=20, b=0), height=300, xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig_spend, use_container_width=True)

    # --- ROW 2 CHARTS ---
    col3, col4, col5 = st.columns(3)
    with col3:
        with st.container(border=True):
            st.markdown("**Channel ROAS**")
            fig_ch_roas = px.bar(pd.DataFrame({'Channel': channels, 'ROAS': [4, 5, 2.5, 3, 1.5, 2]}), x='Channel', y='ROAS')
            fig_ch_roas.update_traces(marker_color='#2a9d8f')
            fig_ch_roas.update_layout(margin=dict(l=0, r=0, t=20, b=0), height=250, xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig_ch_roas, use_container_width=True)

    with col4:
        with st.container(border=True):
            st.markdown("**Budget Allocation**")
            fig_donut = px.pie(spend_data, values='Spend', names='Channel', hole=0.6, color='Channel', color_discrete_map=color_map)
            fig_donut.update_layout(showlegend=False, margin=dict(l=0, r=0, t=20, b=0), height=250)
            st.plotly_chart(fig_donut, use_container_width=True)

    with col5:
        with st.container(border=True):
            st.markdown("**Monthly Spend**")
            spend_trend = pd.DataFrame({'Month': months, 'Spend': [100, 110, 105, 130, 120, 140, 135, 145, 140, 150, 160, 190]})
            fig_spend_trend = px.line(spend_trend, x='Month', y='Spend')
            fig_spend_trend.update_traces(line_color='#0077b6')
            fig_spend_trend.update_layout(margin=dict(l=0, r=0, t=20, b=0), height=250, xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig_spend_trend, use_container_width=True)

# --- OPERATIONS PAGE ---
elif st.session_state.active_page == "Operations":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="op_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="op_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="op_date", label_visibility="collapsed")

    st.title("Operations Dashboard")
    st.markdown("Real-time campaign delivery metrics for Acme Corp")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4, m5, m6 = st.columns(6)
    with m1:
        with st.container(border=True):
            st.metric(label="Impressions", value="49.4M", delta="6.2%")
    with m2:
        with st.container(border=True):
            st.metric(label="Clicks", value="882K", delta="4.8%")
    with m3:
        with st.container(border=True):
            st.metric(label="CPM", value="$6.54", delta="-2.1%")
    with m4:
        with st.container(border=True):
            st.metric(label="CPC", value="$0.37", delta="-3.5%")
    with m5:
        with st.container(border=True):
            st.metric(label="CPA", value="$22.11", delta="-1.8%")
    with m6:
        with st.container(border=True):
            st.metric(label="Pacing", value="82%", delta="1.2%")

    # --- DAILY TRENDS (AREA CHARTS) ---
    days = [f"Mar {i}" for i in range(1, 15)]
    imp_data = pd.DataFrame({'Day': days, 'Impressions': [2.1, 2.3, 2.0, 2.4, 2.6, 2.4, 2.7, 2.3, 2.5, 2.9, 2.6, 2.8, 3.1, 3.0]})
    clicks_data = pd.DataFrame({'Day': days, 'Clicks': [40, 45, 38, 48, 52, 47, 54, 45, 50, 56, 48, 55, 60, 58]})

    col1, col2 = st.columns(2)
    with col1:
        with st.container(border=True):
            st.markdown("**Daily Impressions**\n\n<span style='color:gray; font-size: 14px;'>Last 14 days</span>", unsafe_allow_html=True)
            fig_imp = px.area(imp_data, x='Day', y='Impressions')
            fig_imp.update_traces(line_color='#00a8e8', fillcolor='rgba(0, 168, 232, 0.1)')
            fig_imp.update_layout(margin=dict(l=0, r=0, t=10, b=0), height=250, xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig_imp, use_container_width=True)

    with col2:
        with st.container(border=True):
            st.markdown("**Daily Clicks**\n\n<span style='color:gray; font-size: 14px;'>Last 14 days</span>", unsafe_allow_html=True)
            fig_clicks = px.area(clicks_data, x='Day', y='Clicks')
            fig_clicks.update_traces(line_color='#06d6a0', fillcolor='rgba(6, 214, 160, 0.1)')
            fig_clicks.update_layout(margin=dict(l=0, r=0, t=10, b=0), height=250, xaxis_title=None, yaxis_title=None)
            st.plotly_chart(fig_clicks, use_container_width=True)

    # --- COST METRICS BY CHANNEL ---
    cpm_data = pd.DataFrame({'Channel': channels, 'CPM': [7.5, 9.2, 6.8, 4.5, 3.2, 5.8]})
    cpc_data = pd.DataFrame({'Channel': channels, 'CPC': [0.42, 0.28, 0.35, 0.55, 0.58, 0.40]})
    cpa_data = pd.DataFrame({'Channel': channels, 'CPA': [38.0, 42.0, 35.0, 45.0, 52.0, 48.0]})

    col3, col4, col5 = st.columns(3)
    with col3:
        with st.container(border=True):
            st.markdown("**CPM by Channel**\n\n<span style='color:gray; font-size: 14px;'>Cost per 1,000 impressions</span>", unsafe_allow_html=True)
            fig_cpm = px.bar(cpm_data, x='Channel', y='CPM', color='Channel', color_discrete_map=color_map)
            fig_cpm.update_layout(showlegend=False, margin=dict(l=0, r=0, t=10, b=0), height=250, xaxis_title=None, yaxis_title=None)
            fig_cpm.update_yaxes(tickprefix="$")
            st.plotly_chart(fig_cpm, use_container_width=True)

    with col4:
        with st.container(border=True):
            st.markdown("**CPC by Channel**\n\n<span style='color:gray; font-size: 14px;'>Cost per click</span>", unsafe_allow_html=True)
            fig_cpc = px.bar(cpc_data, x='Channel', y='CPC', color='Channel', color_discrete_map=color_map)
            fig_cpc.update_layout(showlegend=False, margin=dict(l=0, r=0, t=10, b=0), height=250, xaxis_title=None, yaxis_title=None)
            fig_cpc.update_yaxes(tickprefix="$")
            st.plotly_chart(fig_cpc, use_container_width=True)

    with col5:
        with st.container(border=True):
            st.markdown("**CPA by Channel**\n\n<span style='color:gray; font-size: 14px;'>Cost per acquisition</span>", unsafe_allow_html=True)
            fig_cpa = px.bar(cpa_data, x='Channel', y='CPA', color='Channel', color_discrete_map=color_map)
            fig_cpa.update_layout(showlegend=False, margin=dict(l=0, r=0, t=10, b=0), height=250, xaxis_title=None, yaxis_title=None)
            fig_cpa.update_yaxes(tickprefix="$")
            st.plotly_chart(fig_cpa, use_container_width=True)

    # --- BUDGET PACING & CTR ---
    with st.container(border=True):
        st.markdown("**Budget Pacing by Channel**\n\n<span style='color:gray; font-size: 14px;'>Budget vs actual spend</span>", unsafe_allow_html=True)
        
        budget_vals = [120, 100, 70, 80, 25, 40]
        actual_vals = [98, 85, 50, 62, 18, 25]
        
        rev_channels = channels[::-1]
        rev_budget = budget_vals[::-1]
        rev_actual = actual_vals[::-1]
        bar_colors = [color_map[c] for c in rev_channels]
        
        fig_pace = go.Figure()
        
        fig_pace.add_trace(go.Bar(
            y=rev_channels, x=rev_budget, orientation='h', 
            name='Budget', marker=dict(color='#e2e8f0'), 
            width=0.6, hoverinfo='skip'
        ))
        fig_pace.add_trace(go.Bar(
            y=rev_channels, x=rev_actual, orientation='h', 
            name='Actual', marker=dict(color=bar_colors), 
            width=0.4
        ))
        
        fig_pace.update_layout(
            barmode='overlay', showlegend=False, 
            margin=dict(l=0, r=0, t=10, b=0), height=300,
            xaxis_title=None, yaxis_title=None
        )
        fig_pace.update_xaxes(tickprefix="$", ticksuffix="K")
        st.plotly_chart(fig_pace, use_container_width=True)

    with st.container(border=True):
        st.markdown("**Click-Through Rate by Channel**\n\n<span style='color:gray; font-size: 14px;'>CTR %</span>", unsafe_allow_html=True)
        ctr_data = pd.DataFrame({'Channel': channels, 'CTR': [1.8, 3.2, 2.4, 0.9, 1.1, 1.5]})
        fig_ctr = px.bar(ctr_data, x='Channel', y='CTR', color='Channel', color_discrete_map=color_map)
        fig_ctr.update_layout(showlegend=False, margin=dict(l=0, r=0, t=10, b=0), height=300, xaxis_title=None, yaxis_title=None)
        fig_ctr.update_yaxes(ticksuffix="%")
        st.plotly_chart(fig_ctr, use_container_width=True)

# --- CAMPAIGN PERFORMANCE PAGE ---
elif st.session_state.active_page == "Campaign Performance":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="camp_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="camp_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="camp_date", label_visibility="collapsed")

    st.title("Campaign Performance")
    st.markdown("Detailed campaign-level analytics across all channels")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        with st.container(border=True):
            st.metric(label="Total Spend", value="$367K", delta="15.2%")
    with m2:
        with st.container(border=True):
            st.metric(label="Avg ROAS", value="4.14x", delta="6.8%")
    with m3:
        with st.container(border=True):
            st.metric(label="Total Conversions", value="11.4K", delta="9.1%")
    with m4:
        with st.container(border=True):
            st.metric(label="Avg CPA", value="$32.19", delta="-7.3%", delta_color="inverse")

    # --- BAR CHART ---
    campaign_data = pd.DataFrame({
        'Campaign': ['Brand Awareness Q4', 'Retargeting - Cart', 'Prospecting - LAL', 'Holiday Push', 'Summer Sale'],
        'Spend': [85, 42, 65, 120, 55],
        'Revenue': [342, 231, 195, 580, 187],
        'ROAS': ['4x', '5.5x', '3x', '4.8x', '3.4x'],
        'Conversions': ['2,800', '1,900', '1,400', '4,200', '1,100']
    })

    with st.container(border=True):
        st.markdown("**Campaign Performance Comparison**\n\n<span style='color:gray; font-size: 14px;'>Revenue & spend by campaign</span>", unsafe_allow_html=True)
        
        melted_df = pd.melt(campaign_data, id_vars=['Campaign'], value_vars=['Spend', 'Revenue'], var_name='Metric', value_name='Amount')
        
        fig_camp = px.bar(
            melted_df, 
            x='Campaign', 
            y='Amount', 
            color='Metric', 
            barmode='group',
            color_discrete_map={'Spend': '#00a8e8', 'Revenue': '#2a9d8f'}
        )
        
        fig_camp.update_layout(
            legend_title_text='',
            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
            margin=dict(l=0, r=0, t=20, b=0),
            height=350,
            xaxis_title=None,
            yaxis_title=None
        )
        fig_camp.update_yaxes(tickprefix="$", ticksuffix="K")
        st.plotly_chart(fig_camp, use_container_width=True)

    # --- DATA TABLE ---
    with st.container(border=True):
        st.markdown("**Campaign Details**")
        
        display_df = campaign_data.copy()
        display_df['Spend'] = display_df['Spend'].apply(lambda x: f"${x}K")
        display_df['Revenue'] = display_df['Revenue'].apply(lambda x: f"${x}K")
        
        st.dataframe(display_df, use_container_width=True, hide_index=True)

# --- ATTRIBUTION & MMM PAGE ---
elif st.session_state.active_page == "Attribution & MMM":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="attr_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="attr_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="attr_date", label_visibility="collapsed")

    st.title("Attribution & MMM")
    st.markdown("Multi-touch attribution and marketing mix modeling")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        with st.container(border=True):
            st.metric(label="Data-Driven ROAS", value="4.5x", delta="+12.1%")
    with m2:
        with st.container(border=True):
            st.metric(label="Incrementality", value="68%", delta="+3.2%")
    with m3:
        with st.container(border=True):
            st.metric(label="Contribution Margin", value="42%", delta="-1.5%")
    with m4:
        with st.container(border=True):
            st.metric(label="Cross-Channel Assists", value="34%", delta="+8.7%")

    # --- CHARTS ROW ---
    col1, col2 = st.columns([1.6, 1])
    
    with col1:
        with st.container(border=True):
            st.markdown("**Attribution Model Comparison**\n\n<span style='color:gray; font-size: 14px;'>% credit by model</span>", unsafe_allow_html=True)
            
            attr_channels = ['Meta', 'Google', 'TikTok', 'Amazon', 'Email', 'Direct']
            attr_data = pd.DataFrame({
                'Channel': attr_channels * 4,
                'Model': ['First Touch']*6 + ['Last Touch']*6 + ['Linear']*6 + ['Data-Driven']*6,
                'Credit': [
                    32, 28, 18, 12, 5, 5,   
                    28, 35, 12, 18, 4, 3,   
                    30, 30, 15, 15, 6, 4,   
                    35, 28, 17, 10, 5, 3    
                ]
            })
            
            attr_color_map = {
                'First Touch': '#2a9d8f', 
                'Last Touch': '#00a8e8',  
                'Linear': '#7209b7',      
                'Data-Driven': '#ffb703'  
            }
            
            fig_attr = px.bar(
                attr_data, 
                x='Channel', 
                y='Credit', 
                color='Model', 
                barmode='group',
                color_discrete_map=attr_color_map
            )
            
            fig_attr.update_layout(
                legend_title_text='',
                legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
                margin=dict(l=0, r=0, t=10, b=0),
                height=400,
                xaxis_title=None,
                yaxis_title=None,
                yaxis=dict(dtick=9) 
            )
            st.plotly_chart(fig_attr, use_container_width=True)

    with col2:
        with st.container(border=True):
            st.markdown("**Channel Contribution Radar**\n\n<span style='color:gray; font-size: 14px;'>Data-driven model</span>", unsafe_allow_html=True)
            
            radar_vals = [35, 28, 17, 10, 5, 3] 
            
            radar_vals_closed = radar_vals + [radar_vals[0]]
            radar_channels_closed = attr_channels + [attr_channels[0]]
            
            fig_radar = go.Figure()
            fig_radar.add_trace(go.Scatterpolar(
                r=radar_vals_closed,
                theta=radar_channels_closed,
                fill='toself',
                line_color='#00a8e8',
                fillcolor='rgba(0, 168, 232, 0.3)',
                hoverinfo="theta+r"
            ))
            
            fig_radar.update_layout(
                polar=dict(
                    radialaxis=dict(visible=True, showticklabels=False, range=[0, 40])
                ),
                showlegend=False,
                margin=dict(l=40, r=40, t=30, b=30),
                height=400
            )
            st.plotly_chart(fig_radar, use_container_width=True)

# --- INCREMENTALITY TESTS PAGE ---
elif st.session_state.active_page == "Incrementality Tests":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="inc_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="inc_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="inc_date", label_visibility="collapsed")

    st.title("Incrementality Tests")
    st.markdown("Geo-lift and holdout experiment results")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        with st.container(border=True):
            st.metric(label="Avg iROAS", value="3.5x", delta="+9.4%")
    with m2:
        with st.container(border=True):
            st.metric(label="Avg Lift", value="11%", delta="+4.2%")
    with m3:
        with st.container(border=True):
            st.metric(label="Active Tests", value="4", delta="0%")
    with m4:
        with st.container(border=True):
            st.metric(label="Avg Confidence", value="91%", delta="+2.1%")

    # --- CHARTS ROW ---
    col1, col2 = st.columns(2)
    
    with col1:
        with st.container(border=True):
            st.markdown("**Test Results – iROAS**\n\n<span style='color:gray; font-size: 14px;'>By experiment</span>", unsafe_allow_html=True)
            
            iroas_data = pd.DataFrame({
                'Experiment': ['Meta Geo-Lift Q3', 'TikTok Holdout', 'Google Brand Lift', 'Amazon ASIN Test'],
                'iROAS': [3.8, 5.2, 2.9, 2.1]
            })
            
            fig_iroas = px.bar(iroas_data, x='Experiment', y='iROAS')
            fig_iroas.update_traces(marker_color='#2a9d8f')
            fig_iroas.update_layout(
                margin=dict(l=0, r=0, t=10, b=0),
                height=300,
                xaxis_title=None,
                yaxis_title=None
            )
            st.plotly_chart(fig_iroas, use_container_width=True)

    with col2:
        with st.container(border=True):
            st.markdown("**Geo Test vs Control**\n\n<span style='color:gray; font-size: 14px;'>Revenue comparison</span>", unsafe_allow_html=True)
            
            geo_data = pd.DataFrame({
                'Group': ['Test – Northeast', 'Control – Southeast', 'Test – West', 'Control – Midwest'],
                'Revenue': [420, 380, 500, 390]
            })
            
            fig_geo = px.bar(geo_data, x='Group', y='Revenue')
            fig_geo.update_traces(marker_color='#00a8e8')
            fig_geo.update_layout(
                margin=dict(l=0, r=0, t=10, b=0),
                height=300,
                xaxis_title=None,
                yaxis_title=None
            )
            fig_geo.update_yaxes(tickprefix="$", ticksuffix="K")
            st.plotly_chart(fig_geo, use_container_width=True)

    # --- DATA TABLE ---
    with st.container(border=True):
        st.markdown("**Test Library**")
        
        test_library_data = pd.DataFrame({
            'Test Name': ['Meta Geo-Lift Q3', 'TikTok Holdout', 'Google Brand Lift', 'Amazon ASIN Test'],
            'iROAS': ['3.8x', '5.2x', '2.9x', '2.1x'],
            'Lift %': ['+12%', '+18%', '+8%', '+6%'],
            'Confidence': ['95%', '92%', '88%', '90%']
        })
        
        st.dataframe(test_library_data, use_container_width=True, hide_index=True)

# --- MARKETING FUNNEL PAGE ---
elif st.session_state.active_page == "Marketing Funnel":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="funnel_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="funnel_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="funnel_date", label_visibility="collapsed")

    st.title("Marketing Funnel")
    st.markdown("Awareness → Conversion funnel analytics")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        with st.container(border=True):
            st.metric(label="Funnel Conversion", value="0.15%", delta="+2.1%")
    with m2:
        with st.container(border=True):
            st.metric(label="Drop-off Rate", value="62%", delta="-4.3%", delta_color="inverse")
    with m3:
        with st.container(border=True):
            st.metric(label="Avg Time to Convert", value="4.2d", delta="-8.1%", delta_color="inverse")
    with m4:
        with st.container(border=True):
            st.metric(label="Touch Points", value="5.3", delta="+1.2%")

    # --- FUNNEL CHART ---
    with st.container(border=True):
        st.markdown("**Conversion Funnel**\n\n<span style='color:gray; font-size: 14px;'>Volume by stage</span>", unsafe_allow_html=True)
        
        funnel_stages = ['Impressions', 'Clicks', 'Site Visits', 'Add to Cart', 'Checkout', 'Purchase']
        funnel_volumes = [13500000, 800000, 400000, 80000, 30000, 18500]
        
        funnel_colors = ['#2a9d8f', '#06d6a0', '#00a8e8', '#7209b7', '#ffb703', '#e63946']
        
        df_funnel = pd.DataFrame({
            'Stage': funnel_stages,
            'Volume': funnel_volumes,
            'Color': funnel_colors
        })
        
        fig_funnel = px.bar(
            df_funnel, 
            x='Volume', 
            y='Stage', 
            orientation='h',
            color='Stage',
            color_discrete_map={k: v for k, v in zip(funnel_stages, funnel_colors)}
        )
        
        fig_funnel.update_layout(
            showlegend=False,
            margin=dict(l=0, r=0, t=10, b=0),
            height=400,
            xaxis_title=None,
            yaxis_title=None,
            yaxis={'categoryorder':'array', 'categoryarray': funnel_stages[::-1]} 
        )
        
        fig_funnel.update_xaxes(
            tickvals=[0, 3500000, 7000000, 10500000, 14000000],
            ticktext=['0', '3.5M', '7.0M', '10.5M', '14.0M']
        )
        
        st.plotly_chart(fig_funnel, use_container_width=True)

# --- RETAIL MEDIA PAGE ---
elif st.session_state.active_page == "Retail Media":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="rm_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="rm_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="rm_date", label_visibility="collapsed")

    st.title("Retail Media")
    st.markdown("Amazon, Walmart, Instacart & Target performance")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        with st.container(border=True):
            st.metric(label="Avg ACOS", value="18.8%", delta="-5.2%", delta_color="inverse")
    with m2:
        with st.container(border=True):
            st.metric(label="Blended ROAS", value="5.45x", delta="+11.3%")
    with m3:
        with st.container(border=True):
            st.metric(label="New-to-Brand", value="34%", delta="+6.8%")
    with m4:
        with st.container(border=True):
            st.metric(label="Impression Share", value="44%", delta="+3.1%")

    # --- COMPARISON CHART ---
    with st.container(border=True):
        st.markdown("**Retailer Comparison**\n\n<span style='color:gray; font-size: 14px;'>ROAS by retailer</span>", unsafe_allow_html=True)
        
        retailers = ['Amazon', 'Walmart', 'Instacart', 'Target']
        
        rm_data = pd.DataFrame({
            'Retailer': retailers * 2,
            'Metric': ['ROAS'] * 4 + ['New-to-Brand %'] * 4,
            'Value': [
                5.5, 4.2, 6.8, 4.8,  
                32, 28, 41, 35       
            ]
        })
        
        rm_color_map = {
            'ROAS': '#2a9d8f',
            'New-to-Brand %': '#00a8e8'
        }
        
        fig_rm = px.bar(
            rm_data, 
            x='Retailer', 
            y='Value', 
            color='Metric', 
            barmode='group',
            color_discrete_map=rm_color_map
        )
        
        fig_rm.update_layout(
            legend_title_text='',
            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
            margin=dict(l=0, r=0, t=10, b=0),
            height=450,
            xaxis_title=None,
            yaxis_title=None,
            yaxis=dict(dtick=15, range=[0, 65]) 
        )
        
        st.plotly_chart(fig_rm, use_container_width=True)

# --- WALLED GARDENS PAGE ---
elif st.session_state.active_page == "Walled Gardens":
    
    # --- TOP HEADER & FILTERS ---
    filter_col1, filter_col2, filter_col3, _ = st.columns([2, 2, 3, 5])
    with filter_col1:
        st.selectbox("Organization", ["Acme Corp", "Globex", "Initech"], key="wg_org", label_visibility="collapsed")
    with filter_col2:
        st.selectbox("Products", ["All Products", "Software", "Hardware"], key="wg_prod", label_visibility="collapsed")
    with filter_col3:
        st.date_input("Date Range", [], key="wg_date", label_visibility="collapsed")

    st.title("Walled Garden Platforms")
    st.markdown("Meta, Google, TikTok, Pinterest & Snap analytics")

    # --- KPI METRICS ROW ---
    m1, m2, m3, m4 = st.columns(4)
    with m1:
        with st.container(border=True):
            st.metric(label="Avg ROAS", value="3.9x", delta="+7.2%")
    with m2:
        with st.container(border=True):
            st.metric(label="Avg CTR", value="1.82%", delta="+4.5%")
    with m3:
        with st.container(border=True):
            st.metric(label="Engagement Rate", value="3.9%", delta="+12.1%")
    with m4:
        with st.container(border=True):
            st.metric(label="Avg Frequency", value="3.6", delta="-2.3%")

    # --- PLATFORM PERFORMANCE CHART ---
    with st.container(border=True):
        st.markdown("**Platform Performance**\n\n<span style='color:gray; font-size: 14px;'>ROAS & Engagement by platform</span>", unsafe_allow_html=True)
        
        platforms = ['Meta', 'Google', 'TikTok', 'Pinterest', 'Snap']
        
        wg_data = pd.DataFrame({
            'Platform': platforms * 2,
            'Metric': ['ROAS'] * 5 + ['Engagement %'] * 5,
            'Value': [
                4.2, 3.8, 5.0, 2.9, 3.4,  # ROAS values
                3.2, 2.8, 5.5, 4.0, 3.8   # Engagement % values
            ]
        })
        
        wg_color_map = {
            'ROAS': '#2a9d8f',
            'Engagement %': '#7209b7'
        }
        
        fig_wg = px.bar(
            wg_data, 
            x='Platform', 
            y='Value', 
            color='Metric', 
            barmode='group',
            color_discrete_map=wg_color_map
        )
        
        fig_wg.update_layout(
            legend_title_text='',
            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
            margin=dict(l=0, r=0, t=10, b=0),
            height=450,
            xaxis_title=None,
            yaxis_title=None,
            yaxis=dict(dtick=2, range=[0, 8.5]) 
        )
        
        st.plotly_chart(fig_wg, use_container_width=True)

# --- PLACEHOLDER FOR OTHER PAGES ---
else:
    st.title(st.session_state.active_page)
    st.info(f"The {st.session_state.active_page} module is currently under construction.")